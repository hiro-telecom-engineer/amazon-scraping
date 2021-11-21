import os
import logging
import time

from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.select import Select

import pprint
import datetime
import openpyxl

logging.basicConfig(level=logging.INFO)


# 認証の情報
AMAZON_EMAIL = "アカウントメールアドレス"
AMAZON_PASSWORD = "アカウントパスワード"

# 購入履歴情報型
g_hist_info_st = {"購入日": "NULL", "金額": "NULL", "購入品": "NULL"}
# 購入品情報型
g_item_info_st = {"品名": "NULL", "URL": "NULL"}
# 全購入品履歴
g_hist_info_all = list()
# 取得年(この年からさかのぼる)
g_year = 2021

chrome_options = webdriver.ChromeOptions()
chrome_options.add_experimental_option(	"excludeSwitches", ['enable-automation'])
driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)


def main():
	global g_year
	log_in()
	Select(driver.find_element_by_id('orderFilter')).select_by_value('year-'+str(g_year))

	time.sleep(1)
	pages_remaining = True
	while pages_remaining:
		try:
			# 金額、購入日時情報取得
			price_data_N, date_data_N = get_ValueAndDate()
			# 商品情報取得
			item_data_N = get_ItemInfo()
			# 情報集約
			aggregation_data(price_data_N, date_data_N, item_data_N)

			# 「次へ」入力
			try:
				next_link = driver.find_element_by_partial_link_text("次へ")
				# next_link = driver.find_elements_by_xpath("//li[@class='a-last']")
				next_link.click()
				time.sleep(1)
			except NoSuchElementException:
				# 次の年度
				try:
					g_year -= 1
					Select(driver.find_element_by_id('orderFilter')).select_by_value('year-'+str(g_year))
				except NoSuchElementException:
					pprint.pprint(g_hist_info_all, width=240)
					break
		except NoSuchElementException:
			break
	# エクセルへ出力
	output_func()
	driver.quit()  # ブラウザーを終了する。


def log_in():
	# 注文履歴のページを開く。
	logging.info('Navigating...')
	driver.get('https://www.amazon.co.jp/gp/css/order-history')
	time.sleep(1)
	# サインインページにリダイレクトされていることを確認する。
	assert 'Amazonサインイン' in driver.title
	# name="signIn" というサインインフォームを埋める。
	# フォームのname属性の値はブラウザーの開発者ツールで確認できる。
	email_input = driver.find_element_by_name('email')
	email_input.send_keys(AMAZON_EMAIL)  # name="email" という入力ボックスを埋める。
	email_input.send_keys(Keys.RETURN)
	time.sleep(1)
	password_input = driver.find_element_by_name('password')
	password_input.send_keys(AMAZON_PASSWORD)  # name="password" という入力ボックスを埋める。
	time.sleep(2)
	# フォームを送信する。
	logging.info('Signing in...')
	password_input.send_keys(Keys.RETURN)
	time.sleep(1)


# 金額、購入日取得
def get_ValueAndDate():
	get_element = driver.find_elements_by_xpath("//div[@class='a-row a-size-base']")
	data 			= [x.text for x in get_element]
	price_data		= [s for s in data if "￥" in s]
	date_data		= [s for s in data if "年" in s and "月" in s and "日" in s]
	price_data_N	= []
	date_data_N		= []
	# 金額取得
	for data in price_data:
		after_data		= data.replace("￥ ", "")
		next_after_data	= after_data.replace(",", "")
		price_data_N.append(next_after_data)
	# 購入日取得
	for data in date_data:
		after_data = data.replace(" ", "")
		next_after_data = after_data.replace(",", "")
		date_data_N.append(next_after_data)
	return price_data_N, date_data_N


# 商品情報取得
def get_ItemInfo():
	item_element = driver.find_elements_by_xpath("//a[@class='a-link-normal']")
	item = [x.text for x in item_element]
	# 商品名取得
	item_data_N = list()
	item_data_Set = []
	item_count = -1
	for data in item:
		if "注文内容を表示" == data:
			if 0 <= item_count:
				item_data_N.append(item_data_Set.copy())
				item_data_Set = []
			item_count += 1
		elif "" != data and "アカウントサービス" != data:
			after_data = data.replace(" ", "")
			next_after_data = after_data.replace(",", "")
			g_item_info_st["品名"] = next_after_data
			g_item_info_st["URL"] = driver.find_element_by_link_text(data).get_attribute('href')
			item_data_Set.append(g_item_info_st.copy())
	item_data_N.append(item_data_Set.copy())
	return item_data_N


# 情報集約
def aggregation_data(price_data_N, date_data_N, item_data_N):
	global g_hist_info_st
	global g_hist_info_all
	for i in range(len(price_data_N)):
		# 購入履歴情報型
		g_hist_info_st["購入日"] = date_data_N[i]
		g_hist_info_st["金額"] = price_data_N[i]
		g_hist_info_st["購入品"] = item_data_N[i]
		g_hist_info_all.append(g_hist_info_st.copy())
		pprint.pprint(g_hist_info_st, width=240)


# エクセル転記
def output_func():
	global g_hist_info_all
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = '購入履歴一覧'
	# ヘッダ入力
	ws["A1"].value = "購入日"
	ws["B1"].value = "金額"
	ws["C1"].value = "品名"
	ws["D1"].value = "URL"
	# 書き出し
	row_cnt = 2
	for hist_inf in g_hist_info_all:
		ws["A" + str(row_cnt)].value = hist_inf["購入日"]
		ws["B" + str(row_cnt)].value = hist_inf["金額"]
		for hist_item_inf in hist_inf["購入品"]:
			ws["C" + str(row_cnt)].value = hist_item_inf["品名"]
			ws["D" + str(row_cnt)].value = hist_item_inf["URL"]
			row_cnt += 1
	# ファイル保存
	now = datetime.datetime.now()
	file_name = 'amazon購入履歴_{}.xlsx'.format(now.strftime('%Y%m%d_%H%M%S'))
	wb.save(file_name)

if __name__ == '__main__':
	main()
